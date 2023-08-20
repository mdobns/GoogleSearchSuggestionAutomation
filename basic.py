from search_class import GoogleSearchSuggestions
import openpyxl

<<<<<<< HEAD
workbook = openpyxl.load_workbook('Excel.xlsx')
=======
workbook = openpyxl.load_workbook('C:\\Users\\mdobn\\Downloads\\Excel (1).xlsx')
>>>>>>> b9a82819aed53ced4ad95c8f8ffe49cf125cfd93

sheet_names = workbook.sheetnames

google_search = GoogleSearchSuggestions()

for sheet_name in sheet_names:
    worksheet = workbook[sheet_name]
    print(f"Processing data in sheet '{sheet_name}':")
<<<<<<< HEAD
    rows = 2

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        search_query = row[2]
=======
    rows = 2 

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        search_query = row[2] 
>>>>>>> b9a82819aed53ced4ad95c8f8ffe49cf125cfd93
        res1, res2 = google_search.get_max_min_suggestions(search_query)
        rows = rows + 1
        worksheet.cell(row=rows, column=4, value=res1)
        worksheet.cell(row=rows, column=5, value=res2)


<<<<<<< HEAD
workbook.save(f"Excel.xlsx")
google_search.close_browser()

print(f"Excel file 'Excel.xlsx' updated with data.")
=======
workbook.save(f"Excel (1).xlsx")
google_search.close_browser()

print(f"Excel file 'Excel (1).xlsx' updated with data.")
>>>>>>> b9a82819aed53ced4ad95c8f8ffe49cf125cfd93
