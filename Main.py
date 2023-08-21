from GoogleSearchSuggestions import GoogleSearchSuggestions
import openpyxl
import datetime

current_day = datetime.datetime.now().strftime("%A")
workbook = openpyxl.load_workbook('Excel.xlsx')
sheet_names = workbook.sheetnames

google_search = GoogleSearchSuggestions()

for sheet_name in sheet_names:
    if sheet_name == current_day:
        worksheet = workbook[sheet_name]
        print(f"Processing data of workbook :'{sheet_name}'")

        rows = 2

        for row in worksheet.iter_rows(min_row=3, values_only=True):
            search_query = row[2] 
            res1, res2 = google_search.get_max_min_suggestions(search_query)
            rows = rows + 1
            worksheet.cell(row=rows, column=4, value=res1)
            worksheet.cell(row=rows, column=5, value=res2)
print(f"Completed")

workbook.save(f"Excel.xlsx")
google_search.close_browser()

print(f"Excel files workbook {current_day} 'Excel.xlsx' updated with data.")



